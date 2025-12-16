import streamlit as st
import openpyxl
import re
import unicodedata
from io import BytesIO

# Cấu hình trang
st.set_page_config(page_title="Tool Phân Khai 2026 (Fix Column Select)", layout="wide")

st.title("🎯 Tool Phân Khai - Phiên bản Fix Chọn Cột")
st.markdown("""
**Cập nhật:**
✅ ** Tự động bỏ qua cột *"Xã phường trước sáp nhập"* để lấy đúng cột *"Xã, phường, đặc khu"*.
✅ ** Xử lý bất đồng bộ font chữ, khoảng trắng để đảm bảo khớp tên Sheet.
""")

# Upload file
uploaded_file = st.file_uploader("Tải lên file Excel mẫu (Mau_ubnd.xlsx)", type=["xlsx"])

def normalize_text(text):
    """Chuẩn hóa văn bản về dạng so sánh được"""
    if text is None:
        return ""
    text = str(text)
    text = unicodedata.normalize('NFC', text) # Font chuẩn
    text = text.lower() # Chữ thường
    text = text.replace('\xa0', ' ').replace('\t', ' ').replace('\n', ' ')
    text = ' '.join(text.split()) # Xóa khoảng trắng thừa
    return text

def extract_criteria_number(header_text):
    """Lấy số chỉ tiêu từ tiêu đề"""
    if not header_text:
        return None
    text = str(header_text).lower()
    # Tìm số sau chữ "chỉ tiêu" hoặc "ct" hoặc "số"
    match = re.search(r"(?:chỉ tiêu|ct)\s*(?:số)?\s*(\d+)", text)
    if match:
        return str(int(match.group(1))) 
    return None

def is_valid_tt(tt_value):
    """Kiểm tra TT có phải là số (1, 2, 3...)"""
    if tt_value is None:
        return False
    s = str(tt_value).strip().replace('.0', '')
    return s.isdigit()

if uploaded_file is not None:
    if st.button("🚀 Chạy xử lý "):
        try:
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            
            data_map = {} 
            raw_names_found = set() 
            criteria_sheet_count = 0
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # =========================================================
            # BƯỚC 1: QUÉT DỮ LIỆU (LOGIC CHỌN CỘT THÔNG MINH)
            # =========================================================
            status_text.text("Đang quét dữ liệu nguồn...")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                header_row_idx = None
                name_col_idx = None
                
                # 1. Tìm dòng header
                for r in range(1, 16): 
                    row_vals = [cell.value for cell in ws[r]]
                    
                    # Tìm cột tên xã nhưng phải LOẠI TRỪ cột mô tả cũ
                    for idx, val in enumerate(row_vals):
                        v_norm = normalize_text(val)
                        
                        # Điều kiện: Phải có chữ "xã/phường"
                        if "xã" in v_norm and ("phường" in v_norm or "đặc khu" in v_norm):
                            # QUAN TRỌNG: Nếu có chữ "trước" hoặc "sáp nhập" hoặc "cũ" -> BỎ QUA
                            if "trước" in v_norm or "sáp nhập" in v_norm or "cũ" in v_norm:
                                continue
                            
                            # Nếu thỏa mãn -> Chọn làm cột tên
                            header_row_idx = r
                            name_col_idx = idx + 1
                            break
                    if header_row_idx:
                        break
                
                # Nếu xác định được đây là sheet tiêu chí
                if header_row_idx and name_col_idx:
                    criteria_sheet_count += 1
                    col_indices_map = {} 
                    
                    # Quét header đa dòng để lấy số chỉ tiêu
                    scan_rows = [header_row_idx, header_row_idx + 1, header_row_idx + 2]
                    for r_idx in scan_rows:
                        if r_idx > ws.max_row: continue
                        for cell in ws[r_idx]:
                            if cell.column > name_col_idx: # Chỉ quét bên phải cột tên
                                crit_num = extract_criteria_number(cell.value)
                                if crit_num:
                                    col_indices_map[cell.column] = crit_num
                    
                    if col_indices_map:
                        # Lấy dữ liệu
                        start_data = header_row_idx + 1
                        for row in ws.iter_rows(min_row=start_data, values_only=True):
                            raw_name = row[name_col_idx - 1]
                            ward_key = normalize_text(raw_name)
                            
                            if not ward_key or "xã" == ward_key or "ubnd" in ward_key:
                                continue
                            
                            raw_names_found.add(str(raw_name)) 
                            
                            if ward_key not in data_map:
                                data_map[ward_key] = {}
                                
                            for col_idx, crit_num in col_indices_map.items():
                                val = row[col_idx - 1]
                                if val is not None:
                                    data_map[ward_key][crit_num] = val

            st.info(f"Đã quét **{criteria_sheet_count}** sheet tiêu chí. Tìm thấy **{len(data_map)}** xã/phường hợp lệ.")

            # =========================================================
            # BƯỚC 2: ĐIỀN VÀO SHEET XÃ
            # =========================================================
            status_text.text("Đang phân bổ dữ liệu...")
            filled_total = 0
            
            total_sheets = len(wb.sheetnames)
            for i, sheet_name in enumerate(wb.sheetnames):
                progress_bar.progress((i + 1) / total_sheets)
                
                sheet_key = normalize_text(sheet_name)
                
                # Kiểm tra khớp tên
                if sheet_key in data_map:
                    ws = wb[sheet_name]
                    ward_data = data_map[sheet_key]
                    
                    # Tìm cột TT và Kế hoạch
                    tt_col = None
                    target_col = None
                    header_r = None
                    
                    for r in range(1, 20):
                        for cell in ws[r]:
                            v = normalize_text(cell.value)
                            if v == 'tt':
                                tt_col = cell.column
                                header_r = r
                            if "kế hoạch" in v and "2026" in v:
                                target_col = cell.column
                        if tt_col and target_col:
                            break
                    
                    if tt_col and target_col:
                        for row in ws.iter_rows(min_row=header_r + 1):
                            tt_cell = row[tt_col - 1]
                            target_cell = row[target_col - 1]
                            
                            if is_valid_tt(tt_cell.value):
                                current_tt = str(int(float(str(tt_cell.value).strip())))
                                if current_tt in ward_data:
                                    target_cell.value = ward_data[current_tt]
                                    filled_total += 1

            progress_bar.progress(100)
            status_text.text("Hoàn tất!")
            
            # =========================================================
            # KẾT QUẢ
            # =========================================================
            if filled_total > 0:
                st.success(f"🎉 XỬ LÝ THÀNH CÔNG! Đã điền **{filled_total}** ô dữ liệu.")
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                st.download_button("📥 Tải File Kết Quả V5", output, "Ket_qua_Phan_Khai_V5.xlsx")
            else:
                st.error("❌ Vẫn chưa điền được dữ liệu. Kiểm tra lại bảng bên dưới:")
                col1, col2 = st.columns(2)
                with col1:
                    st.warning("Tên tìm thấy (Hy vọng là 'Phường An Đông' thay vì 'Các phường...')")
                    st.write(list(raw_names_found)[:10])
                with col2:
                    st.warning("Tên Sheet đích")
                    st.write(wb.sheetnames[:10])

        except Exception as e:
            st.error(f"Lỗi hệ thống: {e}")